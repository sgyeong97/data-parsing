# -*- coding: utf-8 -*-
"""
Created on Mon Nov  4 14:28:34 2019

@author: kth
"""

import os
import cv2
from PIL import Image
import shutil
from openpyxl import load_workbook

##돌리기 전에 인덱싱 먼저 정확하게 해야함.  indexing_by_matching_table 코드 참고할것.



#원본 동영상 있는 폴더.
rootpath = r"D:\신지영_업무\Bset\Video\0715_Summer"

#이미지 저장할 폴더.
targetpath = r"D:\신지영_업무\Bset\Video\0715_Summer\Images"


matchingTableCheck = False;


imageNumber = 150 ## 몇장씩 뽑을건지.


mode = "jpg"
# "jpg "or "png"



if mode == "jpg":
    imageDir = "JPEGImages"
else:
    imageDir = "PNGImages"

#-------------------------------------------



matchingTablepath = r"D:\신지영_업무\Bset\0122__matchingTable\matchingTable(0721).xlsx"

load_wb = load_workbook(matchingTablepath, data_only=True)
load_ws = load_wb['matching table']
row = 3
index_column = 2
address_column = 3
indice = []
addresses = []
add_row = row


while True:   
    if load_ws.cell(row,index_column).value == None:
        break
    indice.append(load_ws.cell(row,index_column).value)
#    print(load_ws.cell(row,index_column).value)
    row += 1


while True:   
    if load_ws.cell(add_row,address_column).value == None:
        break
    addresses.append(load_ws.cell(add_row,address_column).value)
#    print(load_ws.cell(add_row,address_column).value)
    add_row += 1



def getIndex(address):
    if address in addresses:
        idx = addresses.index(address)
        index = indice[idx]
        return str(index)
    else:
        return address


#-------------------------------------------


def naming(length, name):
    if int(length) == 5:
        
        if len(name) == 1:
            return "0000"+name
        elif len(name) == 2:
            return "000"+name
        elif len(name) == 3:
            return "00" + name
        elif len(name) == 4:
            return "0" + name
        else:
            return name
    elif int(length) == 4:
        if len(name) == 1:
            return "000"+name
        elif len(name) == 2:
            return "00"+name
        elif len(name) == 3:
            return "0" + name
        else:
            return name
        
        
    else: ##6의 경우
        if len(name) == 1:
            return "00000"+name
        elif len(name) == 2:
            return "0000"+name
        elif len(name) == 3:
            return "000" + name
        elif len(name) == 4:
            return "00" + name
        elif len(name) == 5:
            return "0" + name
        else:
            return name


def sampling():
    if not os.path.isdir(targetpath):
        os.mkdir(targetpath)
        
    for (path, dir, files) in os.walk(rootpath):
        for filename in files:
            fname, ext = os.path.splitext(filename)
            if ext == ".mp4":
                weather = os.path.join(rootpath, path).split("\\")[-1].strip()
                if weather == "day":
                    viewState = "d"
                elif weather == "night":
                    viewState = "n"
                elif weather == "rain":
                    viewState = "r"
                elif weather == "snow":
                    viewState = "s"
                else:##분류없음. 이런 경우는 환경별로 분류 안한거고, rootpath 아래에 동영상 무더기로 놓음.
                    viewState = "x"
                
#                index = os.path.abspath(os.path.join(rootpath, path, "..")).split("\\")[-1][:4] #초기 일괄분류때썼던거, 일단안지웠는데 아마 절대 안쓸거임.
                
                address = filename.split("-2020")[0]

                if fname[fname.rfind("_")-1].isdigit():  
                    day = fname[fname.rfind("_")-10:fname.rfind("_")].strip()
                    time = fname.split("_")[-1].split(".")[0].strip()
                else:
                    newline = fname[:fname.rfind("_")]
                    day = newline[newline.rfind("_")-10:newline.rfind("_")].strip()
                    time = newline.split("_")[-1].split(".")[0].strip()
                
                time = time[:-5]  ##000ms제거.
                day = day.replace("-", "")
                time = time.replace("h","")
                time = time.replace("min","")
                time = time.replace("s","")
                
                timeState = day+"_"+time
                
                
                index = getIndex(address)
                
                addname = index + "_" + viewState + "_" + timeState
                if matchingTableCheck:
                    print(addname)
                
                if not matchingTableCheck:
                    if not os.path.isdir(os.path.join(targetpath, addname)):
                        os.mkdir(os.path.join(targetpath, addname))
                
    #                옛버전코드. os.mkdir(os.path.join(targetpath, addname, imageDir))
                    
                    shutil.copy(os.path.join(path, filename), "temp.mp4")
                    
                    cap = cv2.VideoCapture("temp.mp4")
                    length = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                    
                    
                    frame = -1
                    filename = 1
                    ratio = length / imageNumber
                    while(cap.isOpened()):
                        ret, im = cap.read()
                        frame = frame+1
                        if not ret:
                            break
                        if not frame%ratio < 1:
                            continue
                        
                        
    #                    옛버전 코드 .cv2.imwrite(os.path.join(targetpath,addname, imageDir , addname +"_" +naming(4,str(filename)) + ".jpg"), im) ##초기버전임.
    #                     cv2.imwrite(os.path.join(targetpath,addname, addname +"_" +naming(4,str(filename)) + ".jpg"), im)
    #                     filename += 1
                        temp = os.path.join(targetpath,addname,addname+'_'+naming(4,str(filename))+'.jpg')
                        ext = os.path.splitext(temp)[1]
                        result, n = cv2.imencode(ext, im, None)
                        if result:
                            with open(temp, mode="w+b") as f:
                                n.tofile(f)
                        filename += 1
                
                    cap.release()
                    os.remove("temp.mp4")
                

sampling()

print("완료!")